from openpyxl import Workbook
from pathlib import Path
import json

# importe Worksheet e Cell para tipar as variáveis
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell

ROOT_PATH = Path(__file__).parent
PATH_TO_DATA_JSON = ROOT_PATH / "data.json"
PATH_TO_SPREADSHEET = ROOT_PATH / "spreadsheet.xlsm"

# com a classe Workbook é possível criar um arquivo de trabalho
workbook: Workbook = Workbook()

# pegando a planilha ativa (A primeira planilha do arquivo)
worksheet: Worksheet = workbook.active  # type: ignore
# mudando o nome da planilha
worksheet.title = "first sheet"

# criando uma nova planilha na workbook
new_ws = workbook.create_sheet("New sheet")
print(workbook.sheetnames)

# pegando dados do arquivo json
data: list[list]
with open(PATH_TO_DATA_JSON, "r", encoding="utf-8") as file:
    data = json.load(file)

# criando as os nomes das colunas
columns_names = ["name", "old", "sex"]
worksheet.append(columns_names)

for row in data["rows"]:
    worksheet.append(row)

workbook.save(PATH_TO_SPREADSHEET)

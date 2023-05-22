from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path

ROOT_PATH = Path(__file__).parent
PATH_TO_SPREADSHEET = ROOT_PATH / "spreadsheet.xlsm"

workbook: Workbook = load_workbook(PATH_TO_SPREADSHEET)

worksheet: Worksheet = workbook.active

print(worksheet["A1":"B2"])
print()

# iterando pelas linhas
for row in worksheet.iter_rows():
    print(row)

print()

# iterando pelas colunas
for col in worksheet.iter_cols():
    print(col)

print()

# iterando por valores
for row in worksheet.values:
    print(row)
